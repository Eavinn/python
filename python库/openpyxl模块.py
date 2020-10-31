from openpyxl import load_workbook
from openpyxl import Workbook

# 旧表
addr = "aa.xlsx"
wb1 = load_workbook(addr)
ws1 = wb1.active

# 新表
wb2 = Workbook()
ws2 = wb2.active

# 追加
for row in ws1.rows:
    row_value = [i.value for i in row]
    ws2.append(row_value)

# 插入空行
ws1.insert_rows(1)
# 赋值
for idx, value in enumerate(['表头1', '表头2', '表头3']):
    ws1.cell(row=1, column=idx+1, value=value)


wb1.save("bb.xlsx")
